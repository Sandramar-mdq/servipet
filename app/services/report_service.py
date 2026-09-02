"""Etapa 10.2 - Reportes exportables PDF (fpdf2) y Excel (openpyxl).

Servicio central de reportes de solo lectura. Cada funcion de datos
devuelve un dict plano consumible por los renderers, por la vista previa
HTML (/page/reportes) y por los tests. No se modifica la BD.
"""

import io
from datetime import date, datetime, timedelta

from fastapi import HTTPException
from sqlalchemy.orm import Session

from app.models.atencion import AtencionHistorial
from app.models.caja import Caja
from app.models.caja_movimiento import CajaMovimiento
from app.models.comercio import Comercio
from app.models.servicio import Servicio
from app.models.venta import Venta


# ── Helper de datos ──────────────────────────────────────────────

def _nombre_comercio(db: Session, comercio_id: int) -> str:
    row = db.query(Comercio.nombre).filter(Comercio.id == comercio_id).first()
    return row[0] if row else "Comercio"


# ── Servicios de datos ─────────────────────────────────────────────

def resumen_caja(
    db: Session,
    comercio_id: int,
    fecha: date | None = None,
    caja_id: int | None = None,
) -> dict:
    """Resumen de Caja Diaria: ingresos, egresos, balance y ventas."""
    if fecha is None:
        fecha = date.today()

    if caja_id is not None:
        caja = (
            db.query(Caja)
            .filter(Caja.id == caja_id, Caja.comercio_id == comercio_id)
            .first()
        )
        if not caja:
            raise HTTPException(status_code=404, detail="Caja no encontrada")
        monto_inicial = caja.monto_inicial
        inicio = caja.fecha_apertura
        fin = caja.fecha_cierre or datetime.utcnow()
        fecha_reporte = caja.fecha_apertura.date()
    else:
        inicio = datetime.combine(fecha, datetime.min.time())
        fin = datetime.combine(fecha, datetime.max.time())
        monto_inicial = 0.0
        caja_dia = (
            db.query(Caja)
            .filter(
                Caja.comercio_id == comercio_id,
                Caja.fecha_apertura >= inicio,
                Caja.fecha_apertura <= fin,
            )
            .order_by(Caja.fecha_apertura.asc())
            .first()
        )
        if caja_dia:
            monto_inicial = caja_dia.monto_inicial
        fecha_reporte = fecha

    movimientos = (
        db.query(CajaMovimiento)
        .join(Caja, Caja.id == CajaMovimiento.caja_id)
        .filter(
            Caja.comercio_id == comercio_id,
            CajaMovimiento.creado_en >= inicio,
            CajaMovimiento.creado_en <= fin,
        )
        .order_by(CajaMovimiento.creado_en.asc())
        .all()
    )
    ingresos = sum(m.monto for m in movimientos if m.tipo == "INGRESO")
    egresos = sum(m.monto for m in movimientos if m.tipo == "EGRESO")

    ventas = (
        db.query(Venta)
        .filter(
            Venta.comercio_id == comercio_id,
            Venta.fecha >= inicio,
            Venta.fecha <= fin,
            Venta.estado == "COBRADA",
        )
        .order_by(Venta.fecha.asc())
        .all()
    )
    facturacion_ventas = sum(v.total for v in ventas)

    return {
        "tipo": "caja",
        "comercio_nombre": _nombre_comercio(db, comercio_id),
        "fecha": fecha_reporte.isoformat(),
        "monto_inicial": monto_inicial,
        "ingresos": ingresos,
        "egresos": egresos,
        "balance": monto_inicial + ingresos - egresos,
        "facturacion_ventas": facturacion_ventas,
        "cantidad_movimientos": len(movimientos),
        "cantidad_ventas": len(ventas),
        "movimientos": [
            {
                "hora": m.creado_en,
                "tipo": m.tipo,
                "descripcion": m.descripcion,
                "monto": m.monto,
            }
            for m in movimientos
        ],
        "ventas": [
            {
                "id": v.id,
                "hora": v.fecha,
                "medio_pago": v.medio_pago,
                "total": v.total,
            }
            for v in ventas
        ],
    }


def reporte_ventas(
    db: Session,
    comercio_id: int,
    desde: date,
    hasta: date,
) -> dict:
    """Reporte de ventas del POS desglosado por producto, servicio,
    medio de pago y fecha."""
    inicio = datetime.combine(desde, datetime.min.time())
    fin = datetime.combine(hasta, datetime.max.time())

    ventas = (
        db.query(Venta)
        .filter(
            Venta.comercio_id == comercio_id,
            Venta.fecha >= inicio,
            Venta.fecha <= fin,
            Venta.estado == "COBRADA",
        )
        .order_by(Venta.fecha.asc())
        .all()
    )

    cantidad = len(ventas)
    subtotal_total = sum(v.subtotal for v in ventas)
    descuento_total = sum(v.descuento for v in ventas)
    total = sum(v.total for v in ventas)
    ticket_promedio = total / cantidad if cantidad else 0.0

    por_producto: dict[str, dict] = {}
    por_servicio: dict[str, dict] = {}
    por_medio_pago: dict[str, dict] = {}
    por_fecha: dict[str, dict] = {}
    servicios_ids: set[int] = set()

    for v in ventas:
        por_medio_pago.setdefault(v.medio_pago, {
            "medio_pago": v.medio_pago, "cantidad": 0, "total": 0.0,
        })
        por_medio_pago[v.medio_pago]["cantidad"] += 1
        por_medio_pago[v.medio_pago]["total"] += v.total

        fk = v.fecha.date().isoformat()
        por_fecha.setdefault(fk, {"fecha": fk, "cantidad": 0, "total": 0.0})
        por_fecha[fk]["cantidad"] += 1
        por_fecha[fk]["total"] += v.total

        for det in v.detalles:
            if det.tipo == "PRODUCTO":
                nombre = det.producto.nombre if det.producto else "Sin producto"
                bucket = por_producto.setdefault(nombre, {
                    "producto_nombre": nombre, "cantidad": 0, "subtotal": 0.0,
                })
                bucket["cantidad"] += det.cantidad
                bucket["subtotal"] += det.subtotal
            elif det.tipo == "SERVICIO":
                if det.servicio_id:
                    servicios_ids.add(det.servicio_id)
                nombre = f"servicio_{det.servicio_id}" if det.servicio_id else "Sin servicio"
                bucket = por_servicio.setdefault(str(det.servicio_id), {
                    "servicio_id": det.servicio_id, "servicio_nombre": nombre,
                    "cantidad": 0, "total": 0.0,
                })
                bucket["cantidad"] += det.cantidad
                bucket["total"] += det.subtotal

    if servicios_ids:
        nombres_servicios = {
            s.id: s.nombre
            for s in db.query(Servicio).filter(Servicio.id.in_(servicios_ids)).all()
        }
        for bucket in por_servicio.values():
            sid = bucket["servicio_id"]
            bucket["servicio_nombre"] = nombres_servicios.get(sid, "Sin servicio")

    return {
        "tipo": "ventas",
        "comercio_nombre": _nombre_comercio(db, comercio_id),
        "desde": desde.isoformat(),
        "hasta": hasta.isoformat(),
        "resumen": {
            "cantidad_ventas": cantidad,
            "subtotal_total": subtotal_total,
            "descuento_total": descuento_total,
            "total": total,
            "ticket_promedio": ticket_promedio,
        },
        "por_producto": sorted(por_producto.values(), key=lambda x: x["subtotal"], reverse=True),
        "por_servicio": sorted(por_servicio.values(), key=lambda x: x["total"], reverse=True),
        "por_medio_pago": sorted(por_medio_pago.values(), key=lambda x: x["total"], reverse=True),
        "por_fecha": sorted(por_fecha.values(), key=lambda x: x["fecha"]),
    }


def reporte_metricas(db: Session, comercio_id: int, dias: int = 30) -> dict:
    """Metricas consolidadas para el administrador."""
    from app.services.dashboard import metricas, resumen_dia

    met = metricas(db, comercio_id, dias)
    hoy = resumen_dia(db, comercio_id)
    desde = datetime.combine(date.today() - timedelta(days=dias), datetime.min.time())

    ventas = (
        db.query(Venta)
        .filter(
            Venta.comercio_id == comercio_id,
            Venta.fecha >= desde,
            Venta.estado == "COBRADA",
        )
        .all()
    )
    atenciones = db.query(AtencionHistorial).filter(AtencionHistorial.fecha >= desde).all()
    total_ventas = sum(v.total for v in ventas)
    total_atenciones = sum(a.monto_cobrado for a in atenciones)

    return {
        "tipo": "metricas",
        "comercio_nombre": _nombre_comercio(db, comercio_id),
        "dias": dias,
        "periodo": met["periodo"],
        "resumen_hoy": hoy,
        "servicios_mas_pedidos": met["servicios_mas_pedidos"],
        "horas_pico": met["horas_pico"],
        "productos_mas_vendidos": met["productos_mas_vendidos"],
        "total_facturado_ventas": total_ventas,
        "total_facturado_atenciones": total_atenciones,
        "total_facturado_periodo": total_ventas + total_atenciones,
        "cantidad_ventas": len(ventas),
        "cantidad_atenciones": len(atenciones),
    }


# ── Helpers de render ─────────────────────────────────────────────

def _money(valor) -> str:
    return f"${float(valor or 0.0):,.2f}"


def _fecha(dt) -> str:
    return dt.strftime("%d/%m/%Y")


def _fecha_hora(dt) -> str:
    return dt.strftime("%d/%m/%Y %H:%M")


def _truncar(texto: object, maximo: int = 40) -> str:
    s = str(texto)
    return s if len(s) <= maximo else s[: maximo - 3] + "..."


PDF_W = 190


def _pdf_inicio(titulo: str, subtitulo: str):
    from fpdf import FPDF

    pdf = FPDF(format="A4")
    pdf.set_margins(10, 10, 10)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("helvetica", "B", 14)
    pdf.cell(PDF_W, 8, titulo, align="C")
    pdf.ln()
    pdf.set_font("helvetica", "", 9)
    pdf.cell(PDF_W, 5, subtitulo, align="C")
    pdf.ln(4)
    return pdf


def _pdf_par(pdf, etiqueta: str, valor: str):
    pdf.set_font("helvetica", "", 9)
    pdf.cell(70, 5, etiqueta)
    pdf.set_font("helvetica", "B", 9)
    pdf.cell(PDF_W - 70, 5, valor, align="R")
    pdf.ln(5)


def _pdf_tabla(pdf, encabezados, filas, anchos, alineaciones=None):
    fila_alto = 5
    pdf.set_font("helvetica", "B", 8)
    pdf.set_fill_color(235, 235, 240)
    pdf.set_text_color(20, 20, 30)
    for i, enc in enumerate(encabezados):
        pdf.cell(anchos[i], fila_alto, _truncar(enc, 30), border=1, fill=True, align="C")
    pdf.ln()
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("helvetica", "", 8)
    pdf.set_fill_color(255, 255, 255)

    def _encabezado_repetido():
        pdf.set_font("helvetica", "B", 8)
        pdf.set_fill_color(235, 235, 240)
        pdf.set_text_color(20, 20, 30)
        for i, enc in enumerate(encabezados):
            pdf.cell(anchos[i], fila_alto, _truncar(enc, 30), border=1, fill=True, align="C")
        pdf.ln()
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("helvetica", "", 8)

    if not filas:
        pdf.set_font("helvetica", "I", 8)
        pdf.cell(PDF_W, fila_alto, "Sin datos", border=1, align="C")
        pdf.ln()
        return

    for fila in filas:
        if pdf.will_page_break(fila_alto):
            pdf.add_page()
            _encabezado_repetido()
        for i, celda in enumerate(fila):
            align = "C" if not alineaciones else alineaciones[i]
            pdf.cell(anchos[i], fila_alto, _truncar(celda, 30), border=1, align=align)
        pdf.ln()


def _pdf_pie(pdf, filas_resumen) -> None:
    pdf.ln(2)
    pdf.set_font("helvetica", "B", 9)
    for etiqueta, valor in filas_resumen:
        pdf.cell(60, 6, etiqueta)
        pdf.cell(PDF_W - 60, 6, valor, align="R")
        pdf.ln(6)


# ── Renderers PDF ─────────────────────────────────────────────────

def pdf_reporte_caja(data: dict) -> bytes:
    pdf = _pdf_inicio(
        "Reporte Caja Diaria",
        f"{data['comercio_nombre']} - {data['fecha']} - Emitido {_fecha_hora(datetime.now())}",
    )
    _pdf_par(pdf, "Monto inicial", _money(data["monto_inicial"]))
    _pdf_par(pdf, "Ingresos", _money(data["ingresos"]))
    _pdf_par(pdf, "Egresos", _money(data["egresos"]))
    _pdf_par(pdf, "Balance", _money(data["balance"]))
    _pdf_par(pdf, "Facturacion ventas", _money(data["facturacion_ventas"]))
    _pdf_par(pdf, "Ventas registradas", str(data["cantidad_ventas"]))
    pdf.ln(3)

    pdf.set_font("helvetica", "B", 10)
    pdf.cell(PDF_W, 6, "Movimientos")
    pdf.ln(6)
    filas = [
        (_fecha_hora(m["hora"]), m["tipo"], m["descripcion"], _money(m["monto"]))
        for m in data["movimientos"]
    ]
    _pdf_tabla(pdf, ["Fecha", "Tipo", "Descripcion", "Monto"], filas, [45, 18, 100, 27], ["L", "C", "L", "R"])

    pdf.ln(2)
    pdf.set_font("helvetica", "B", 10)
    pdf.cell(PDF_W, 6, "Ventas del dia")
    pdf.ln(6)
    filas_v = [
        (str(v["id"]), _fecha_hora(v["hora"]), v["medio_pago"], _money(v["total"]))
        for v in data["ventas"]
    ]
    _pdf_tabla(pdf, ["Venta", "Fecha", "Medio de pago", "Total"], filas_v, [20, 50, 40, 27], ["C", "L", "L", "R"])
    return bytes(pdf.output())


def pdf_reporte_ventas(data: dict) -> bytes:
    r = data["resumen"]
    pdf = _pdf_inicio(
        "Reporte de Ventas",
        f"{data['comercio_nombre']} - {data['desde']} a {data['hasta']} - Emitido {_fecha_hora(datetime.now())}",
    )
    _pdf_par(pdf, "Cantidad de ventas", str(r["cantidad_ventas"]))
    _pdf_par(pdf, "Subtotal", _money(r["subtotal_total"]))
    _pdf_par(pdf, "Descuentos", _money(r["descuento_total"]))
    _pdf_par(pdf, "Total", _money(r["total"]))
    _pdf_par(pdf, "Ticket promedio", _money(r["ticket_promedio"]))
    pdf.ln(3)

    pdf.set_font("helvetica", "B", 10)
    pdf.cell(PDF_W, 6, "Desglose por producto")
    pdf.ln(6)
    _pdf_tabla(
        pdf, ["Producto", "Cantidad", "Subtotal"],
        [(p["producto_nombre"], str(p["cantidad"]), _money(p["subtotal"])) for p in data["por_producto"]],
        [120, 30, 40], ["L", "C", "R"],
    )

    pdf.ln(2)
    pdf.set_font("helvetica", "B", 10)
    pdf.cell(PDF_W, 6, "Desglose por servicio")
    pdf.ln(6)
    _pdf_tabla(
        pdf, ["Servicio", "Cantidad", "Total"],
        [(s["servicio_nombre"], str(s["cantidad"]), _money(s["total"])) for s in data["por_servicio"]],
        [120, 30, 40], ["L", "C", "R"],
    )

    pdf.ln(2)
    pdf.set_font("helvetica", "B", 10)
    pdf.cell(PDF_W, 6, "Desglose por medio de pago")
    pdf.ln(6)
    _pdf_tabla(
        pdf, ["Medio de pago", "Cantidad", "Total"],
        [(mp["medio_pago"], str(mp["cantidad"]), _money(mp["total"])) for mp in data["por_medio_pago"]],
        [120, 30, 40], ["L", "C", "R"],
    )

    pdf.ln(2)
    pdf.set_font("helvetica", "B", 10)
    pdf.cell(PDF_W, 6, "Desglose por fecha")
    pdf.ln(6)
    _pdf_tabla(
        pdf, ["Fecha", "Cantidad", "Total"],
        [(pf["fecha"], str(pf["cantidad"]), _money(pf["total"])) for pf in data["por_fecha"]],
        [120, 30, 40], ["L", "C", "R"],
    )
    return bytes(pdf.output())


def pdf_reporte_metricas(data: dict) -> bytes:
    hoy = data["resumen_hoy"]
    pdf = _pdf_inicio(
        "Metricas y Estadisticas",
        f"{data['comercio_nombre']} - {data['periodo']} - Emitido {_fecha_hora(datetime.now())}",
    )
    _pdf_par(pdf, "Ventas hoy", _money(hoy["facturacion_total"]))
    _pdf_par(pdf, "Atenciones hoy", str(hoy["cantidad_atenciones"]))
    _pdf_par(pdf, "Total facturado (periodo)", _money(data["total_facturado_periodo"]))
    _pdf_par(pdf, "Ventas periodo", _money(data["total_facturado_ventas"]))
    _pdf_par(pdf, "Atenciones periodo", _money(data["total_facturado_atenciones"]))
    _pdf_par(pdf, "Ventas registradas", str(data["cantidad_ventas"]))
    pdf.ln(3)

    pdf.set_font("helvetica", "B", 10)
    pdf.cell(PDF_W, 6, "Servicios mas pedidos")
    pdf.ln(6)
    _pdf_tabla(
        pdf, ["Servicio", "Cantidad"],
        [(s["servicio_nombre"], str(s["cantidad"])) for s in data["servicios_mas_pedidos"]],
        [160, 30], ["L", "C"],
    )

    pdf.ln(2)
    pdf.set_font("helvetica", "B", 10)
    pdf.cell(PDF_W, 6, "Horas pico")
    pdf.ln(6)
    _pdf_tabla(
        pdf, ["Hora", "Turnos"],
        [(f"{h['hora']:02d}:00", str(h["cantidad_turnos"])) for h in data["horas_pico"]],
        [160, 30], ["L", "C"],
    )

    pdf.ln(2)
    pdf.set_font("helvetica", "B", 10)
    pdf.cell(PDF_W, 6, "Productos mas vendidos")
    pdf.ln(6)
    _pdf_tabla(
        pdf, ["Producto", "Cantidad", "Total"],
        [
            (p["producto_nombre"], str(p["cantidad_vendida"]), _money(p["total_facturado"]))
            for p in data["productos_mas_vendidos"]
        ],
        [110, 30, 50], ["L", "C", "R"],
    )
    return bytes(pdf.output())


# ── Renderers Excel ───────────────────────────────────────────────

def _wb_base(titulo: str, ws, data: dict) -> None:
    from openpyxl.styles import Font

    ws.append([titulo])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([data["comercio_nombre"]])
    ws.append([data.get("fecha") or f"{data.get('desde', '')} a {data.get('hasta', '')}"])
    ws.append([])


def _wb_tabla(ws, encabezados: list[str], filas: list[list]) -> None:
    from openpyxl.styles import Font

    ws.append(encabezados)
    for celda in ws[ws.max_row]:
        celda.font = Font(bold=True)
    for fila in filas:
        ws.append(list(fila))
    for col_idx in range(1, len(encabezados) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 28


def _wb_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def xlsx_reporte_caja(data: dict) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    _wb_base("Reporte Caja Diaria", ws, data)
    ws.append(["Monto inicial", data["monto_inicial"]])
    ws.append(["Ingresos", data["ingresos"]])
    ws.append(["Egresos", data["egresos"]])
    ws.append(["Balance", data["balance"]])
    ws.append(["Facturacion ventas", data["facturacion_ventas"]])
    ws.append(["Ventas registradas", data["cantidad_ventas"]])

    ws_mov = wb.create_sheet("Movimientos")
    _wb_tabla(ws_mov, ["Fecha", "Tipo", "Descripcion", "Monto"], [
        [_fecha_hora(m["hora"]), m["tipo"], m["descripcion"], m["monto"]]
        for m in data["movimientos"]
    ])

    ws_ventas = wb.create_sheet("Ventas")
    _wb_tabla(ws_ventas, ["Venta", "Fecha", "Medio de pago", "Total"], [
        [v["id"], _fecha_hora(v["hora"]), v["medio_pago"], v["total"]]
        for v in data["ventas"]
    ])
    return _wb_bytes(wb)


def xlsx_reporte_ventas(data: dict) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    r = data["resumen"]
    _wb_base("Reporte de Ventas", ws, data)
    ws.append(["Cantidad de ventas", r["cantidad_ventas"]])
    ws.append(["Subtotal", r["subtotal_total"]])
    ws.append(["Descuentos", r["descuento_total"]])
    ws.append(["Total", r["total"]])
    ws.append(["Ticket promedio", r["ticket_promedio"]])

    ws_prod = wb.create_sheet("PorProducto")
    _wb_tabla(ws_prod, ["Producto", "Cantidad", "Subtotal"], [
        [p["producto_nombre"], p["cantidad"], p["subtotal"]] for p in data["por_producto"]
    ])

    ws_svc = wb.create_sheet("PorServicio")
    _wb_tabla(ws_svc, ["Servicio", "Cantidad", "Total"], [
        [s["servicio_nombre"], s["cantidad"], s["total"]] for s in data["por_servicio"]
    ])

    ws_mp = wb.create_sheet("PorMedioPago")
    _wb_tabla(ws_mp, ["Medio de pago", "Cantidad", "Total"], [
        [mp["medio_pago"], mp["cantidad"], mp["total"]] for mp in data["por_medio_pago"]
    ])

    ws_fec = wb.create_sheet("PorFecha")
    _wb_tabla(ws_fec, ["Fecha", "Cantidad", "Total"], [
        [pf["fecha"], pf["cantidad"], pf["total"]] for pf in data["por_fecha"]
    ])
    return _wb_bytes(wb)


def xlsx_reporte_metricas(data: dict) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    _wb_base("Metricas y Estadisticas", ws, data)
    ws.append(["Periodo", data["periodo"]])
    ws.append(["Total facturado (periodo)", data["total_facturado_periodo"]])
    ws.append(["Ventas periodo", data["total_facturado_ventas"]])
    ws.append(["Atenciones periodo", data["total_facturado_atenciones"]])

    ws_svc = wb.create_sheet("Servicios")
    _wb_tabla(ws_svc, ["Servicio", "Cantidad"], [
        [s["servicio_nombre"], s["cantidad"]] for s in data["servicios_mas_pedidos"]
    ])

    ws_h = wb.create_sheet("HorasPico")
    _wb_tabla(ws_h, ["Hora", "Turnos"], [
        [f"{h['hora']:02d}:00", h["cantidad_turnos"]] for h in data["horas_pico"]
    ])

    ws_prod = wb.create_sheet("Productos")
    _wb_tabla(ws_prod, ["Producto", "Cantidad", "Total"], [
        [p["producto_nombre"], p["cantidad_vendida"], p["total_facturado"]]
        for p in data["productos_mas_vendidos"]
    ])
    return _wb_bytes(wb)
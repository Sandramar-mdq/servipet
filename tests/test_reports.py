import io
from datetime import date, timedelta

from openpyxl import load_workbook

from tests.conftest import TestingSessionLocal

XLSX_MT = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PDF_MT = "application/pdf"


class TestReportesCaja:
    def _setup(self, client, admin_headers):
        r = client.post("/productos/", headers=admin_headers, json={
            "nombre": "Producto Reporte",
            "precio_venta": 1500.0,
            "stock_actual": 10,
        })
        pid = r.json()["id"]
        client.post("/caja/abrir", headers=admin_headers, json={"monto_inicial": 10000.0})
        client.post("/caja/movimiento", headers=admin_headers, json={
            "tipo": "INGRESO", "monto": 3000.0, "descripcion": "Venta 1",
        })
        client.post("/caja/movimiento", headers=admin_headers, json={
            "tipo": "EGRESO", "monto": 500.0, "descripcion": "Insumos",
        })
        resp = client.post("/ventas/", headers=admin_headers, json={
            "medio_pago": "debito",
            "detalles": [{"tipo": "PRODUCTO", "producto_id": pid, "cantidad": 2, "precio_unitario": 1500.0}],
        })
        assert resp.status_code == 201

    def test_service_resumen_caja(self, client, admin_headers):
        self._setup(client, admin_headers)
        from app.services.report_service import resumen_caja

        db = TestingSessionLocal()
        try:
            data = resumen_caja(db, comercio_id=1, fecha=date.today())
        finally:
            db.close()
        assert data["monto_inicial"] == 10000.0
        assert data["ingresos"] == 3000.0
        assert data["egresos"] == 500.0
        assert data["balance"] == 12500.0
        assert data["facturacion_ventas"] == 3000.0
        assert data["cantidad_movimientos"] == 2
        assert data["cantidad_ventas"] == 1

    def test_pdf_caja(self, client, admin_headers):
        self._setup(client, admin_headers)
        resp = client.get("/reportes/caja/pdf", headers=admin_headers)
        assert resp.status_code == 200
        assert resp.headers["content-type"] == PDF_MT
        assert "filename=" in resp.headers["content-disposition"]
        assert resp.content.startswith(b"%PDF")
        assert len(resp.content) > 200

    def test_pdf_caja_fecha(self, client, admin_headers):
        self._setup(client, admin_headers)
        resp = client.get("/reportes/caja/pdf?fecha=2026-01-15", headers=admin_headers)
        assert resp.status_code == 200
        assert resp.content.startswith(b"%PDF")
        assert "reporte_caja_2026-01-15.pdf" in resp.headers["content-disposition"]

    def test_pdf_caja_no_existe_404(self, client, admin_headers):
        resp = client.get("/reportes/caja/pdf?caja_id=999", headers=admin_headers)
        assert resp.status_code == 404

    def test_excel_caja(self, client, admin_headers):
        self._setup(client, admin_headers)
        resp = client.get("/reportes/caja/excel", headers=admin_headers)
        assert resp.status_code == 200
        assert resp.headers["content-type"] == XLSX_MT
        assert resp.content[:2] == b"PK"
        wb = load_workbook(io.BytesIO(resp.content))
        assert wb.sheetnames == ["Resumen", "Movimientos", "Ventas"]
        ws = wb["Resumen"]
        filas = {r[0]: r[1] for r in ws.iter_rows(min_row=5, values_only=True) if r[0]}
        assert filas["Monto inicial"] == 10000.0
        assert filas["Balance"] == 12500.0


class TestReportesVentas:
    def _setup(self, client, admin_headers):
        r = client.post("/productos/", headers=admin_headers, json={
            "nombre": "Producto Reporte",
            "precio_venta": 1500.0,
            "stock_actual": 10,
        })
        pid = r.json()["id"]
        resp = client.post("/ventas/", headers=admin_headers, json={
            "medio_pago": "debito",
            "detalles": [{"tipo": "PRODUCTO", "producto_id": pid, "cantidad": 2, "precio_unitario": 1500.0}],
        })
        assert resp.status_code == 201

    def test_service_reporte_ventas(self, client, admin_headers):
        self._setup(client, admin_headers)
        from app.services.report_service import reporte_ventas

        db = TestingSessionLocal()
        try:
            data = reporte_ventas(db, comercio_id=1, desde=date.today() - timedelta(days=30), hasta=date.today())
        finally:
            db.close()
        assert data["resumen"]["cantidad_ventas"] == 1
        assert data["resumen"]["total"] == 3000.0
        assert data["resumen"]["ticket_promedio"] == 3000.0
        assert data["por_producto"][0]["producto_nombre"] == "Producto Reporte"
        assert data["por_producto"][0]["cantidad"] == 2
        assert data["por_producto"][0]["subtotal"] == 3000.0
        assert data["por_medio_pago"][0]["medio_pago"] == "debito"
        assert data["por_medio_pago"][0]["total"] == 3000.0
        assert data["por_fecha"][0]["fecha"] == date.today().isoformat()

    def test_pdf_ventas(self, client, admin_headers):
        self._setup(client, admin_headers)
        resp = client.get("/reportes/ventas/pdf", headers=admin_headers)
        assert resp.status_code == 200
        assert resp.headers["content-type"] == PDF_MT
        assert resp.content.startswith(b"%PDF")

    def test_excel_ventas(self, client, admin_headers):
        self._setup(client, admin_headers)
        resp = client.get("/reportes/ventas/excel", headers=admin_headers)
        assert resp.status_code == 200
        assert resp.content[:2] == b"PK"
        wb = load_workbook(io.BytesIO(resp.content))
        assert wb.sheetnames == ["Resumen", "PorProducto", "PorServicio", "PorMedioPago", "PorFecha"]
        ws = wb["PorProducto"]
        assert ws["A2"].value == "Producto Reporte"
        assert ws["B2"].value == 2
        assert ws["C2"].value == 3000.0


class TestReportesMetricas:
    def test_service_reporte_metricas(self, client, admin_headers):
        from app.services.report_service import reporte_metricas

        db = TestingSessionLocal()
        try:
            data = reporte_metricas(db, comercio_id=1, dias=7)
        finally:
            db.close()
        assert data["periodo"] == "ultimos 7 dias"
        assert data["total_facturado_periodo"] == 0.0
        assert data["servicios_mas_pedidos"] == []
        assert data["horas_pico"] == []
        assert data["productos_mas_vendidos"] == []

    def test_pdf_metricas(self, client, admin_headers):
        resp = client.get("/reportes/metricas/pdf?dias=7", headers=admin_headers)
        assert resp.status_code == 200
        assert resp.headers["content-type"] == PDF_MT
        assert resp.content.startswith(b"%PDF")

    def test_pdf_metricas_sin_datos(self, client, admin_headers):
        resp = client.get("/reportes/metricas/pdf", headers=admin_headers)
        assert resp.status_code == 200
        assert resp.content.startswith(b"%PDF")
        assert len(resp.content) > 200

    def test_excel_metricas(self, client, admin_headers):
        resp = client.get("/reportes/metricas/excel?dias=7", headers=admin_headers)
        assert resp.status_code == 200
        assert resp.content[:2] == b"PK"
        wb = load_workbook(io.BytesIO(resp.content))
        assert wb.sheetnames == ["Resumen", "Servicios", "HorasPico", "Productos"]


class TestReportesAuthYPagina:
    def test_no_auth_pdf(self, client):
        resp = client.get("/reportes/caja/pdf")
        assert resp.status_code == 401

    def test_no_auth_excel(self, client):
        resp = client.get("/reportes/ventas/excel")
        assert resp.status_code == 401

    def test_page_reportes(self, client):
        resp = client.get("/page/reportes")
        assert resp.status_code == 200
        assert "Reportes" in resp.text
        assert "Descargar PDF" in resp.text

    def test_page_reportes_ventas(self, client):
        resp = client.get("/page/reportes?tipo=ventas")
        assert resp.status_code == 200
        assert "Desglose por producto" in resp.text

    def test_page_reportes_metricas(self, client):
        resp = client.get("/page/reportes?tipo=metricas&dias=7")
        assert resp.status_code == 200
        assert "Total facturado" in resp.text